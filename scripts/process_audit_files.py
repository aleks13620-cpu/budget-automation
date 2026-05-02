"""
Обрабатывает 12 аудит-файлов ФАЗА 0 из data/uploads/
→ сохраняет по одному JSON в scripts/benchmark-ready/
Имя выходного файла: {поставщик}_{итоговая_сумма}.json

Ключевая строка = строка, где col C (поставщик) И col E (наименования) заполнены.
Строки только с col B — игнорируются.
"""
import openpyxl
import json
import re
from pathlib import Path

UPLOADS_DIR = Path(__file__).parent.parent / "data" / "uploads"
OUTPUT_DIR = Path(__file__).parent / "benchmark-ready"


def split_semicolons(text):
    if not text:
        return []
    parts = str(text).split(';')
    result = []
    for p in parts:
        cleaned = re.sub(r'[\n"]+', ' ', p).strip()
        if cleaned:
            result.append(cleaned)
    return result


def parse_price(s):
    s = str(s).strip().replace(',', '.').replace('\xa0', '').replace(' ', '')
    try:
        return float(s)
    except ValueError:
        return None


def safe_filename(s):
    s = re.sub(r'[<>:"/\\|?*\n\r]', '', str(s)).strip()
    return s[:80]


def process_file(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        ws = wb.active

        # Найти ключевую строку: col 3 (поставщик) И col 5 (наименования) не пусты
        key_row = None
        for r in range(4, ws.max_row + 1):
            supplier = ws.cell(r, 3).value
            names_val = ws.cell(r, 5).value
            if supplier and names_val:
                key_row = r
                break

        if key_row is None:
            return None, "ключевая строка не найдена"

        def cell(c):
            return ws.cell(key_row, c).value

        source_invoice = str(cell(2) or "")
        supplier      = str(cell(3) or "")
        pos_count_raw = cell(4)
        names_raw     = cell(5)
        articles_raw  = cell(6)
        name_status   = str(cell(7) or "")
        prices_raw    = cell(8)
        price_status  = str(cell(9) or "")
        qty_raw       = cell(10)
        qty_status    = str(cell(11) or "")
        comment       = str(cell(12) or "")

        names     = split_semicolons(names_raw)
        articles  = split_semicolons(articles_raw)
        prices    = [parse_price(p) for p in split_semicolons(prices_raw)]
        qtys      = split_semicolons(qty_raw)

        warnings = []
        if prices and len(names) != len(prices):
            warnings.append(f"длины не совпадают: имён={len(names)}, цен={len(prices)}")

        items = []
        for i, name in enumerate(names):
            price = prices[i] if i < len(prices) else None
            # Баг 2: позиции с null-ценой помечаются явно и попадают в warnings
            if price is None:
                warnings.append(f"нет цены у позиции {i+1}: «{name[:60]}»")
            items.append({
                "item_index":    i,
                "name":          name,
                "article":       articles[i] if i < len(articles) else "",
                "price_with_vat": price,
                "quantity":      qtys[i] if i < len(qtys) else "",
            })

        total_sum = round(sum(p for p in prices if p is not None), 2)
        position_count = int(pos_count_raw) if pos_count_raw else len(items)

        result = {
            "audit_file":      xlsx_path.name,
            "source_invoice":  source_invoice,
            "supplier":        supplier,
            "position_count":  position_count,
            "total_sum":       total_sum,
            "name_status":     name_status,
            "price_status":    price_status,
            "qty_status":      qty_status,
            "comment":         comment,
            "warnings":        warnings,
            "items":           items,
        }
        return result, None
    finally:
        wb.close()  # Баг 3: закрываем workbook в любом случае


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)
    xlsx_files = sorted(UPLOADS_DIR.glob("*.xlsx"))

    processed, skipped = 0, 0

    for path in xlsx_files:
        print(f"[>] {path.name}")
        result, err = process_file(path)

        if result is None:
            print(f"   PROPUSK: {err}")
            skipped += 1
            continue

        supplier_clean = safe_filename(result["supplier"])
        total_str = f"{result['total_sum']:.2f}".replace('.', '_')
        base = f"{supplier_clean}_{total_str}" if supplier_clean else path.stem
        out_path = OUTPUT_DIR / f"{base}.json"
        # Баг 1: защита от коллизии — добавляем суффикс вместо перезаписи
        n = 2
        while out_path.exists():
            out_path = OUTPUT_DIR / f"{base}_{n}.json"
            n += 1

        with open(out_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)

        warn_str = f"  WARN: {', '.join(result['warnings'])}" if result['warnings'] else ""
        print(f"   OK  {out_path.name}  ({len(result['items'])} items, sum={result['total_sum']}){warn_str}")
        processed += 1

    print(f"\nИтого: обработано={processed}, пропущено={skipped}")
    print(f"Папка: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
