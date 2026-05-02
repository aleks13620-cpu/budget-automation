"""Create phase0_sample.xlsx — образец правильного заполнения эталонной таблицы."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── данные строки 6 (радиаторы ИТАС) ──────────────────────────────────────────
ITEMS = [
    ("Радиатор стальной EVRA Ventil Hygiene HV 20-500-400 без крепежа",   "HV 20-500-400",  "4067,48",  "6 шт"),
    ("Радиатор стальной EVRA Ventil Hygiene HV 20-500-900 без крепежа",   "HV 20-500-900",  "6305,95",  "1 шт"),
    ("Радиатор стальной EVRA Ventil Hygiene HV 20-500-1000 без крепежа",  "HV 20-500-1000", "6901,26",  "2 шт"),
    ("Радиатор стальной EVRA Ventil Hygiene HV 20-500-1100 без крепежа",  "HV 20-500-1100", "7352,39",  "3 шт"),
    ("Радиатор стальной EVRA Ventil Hygiene HV 20-500-1200 без крепежа",  "HV 20-500-1200", "7788,01",  "8 шт"),
    ("Радиатор стальной EVRA Ventil Hygiene HV 20-500-1600 без крепежа",  "HV 20-500-1600", "9675,21",  "6 шт"),
    ("Радиатор стальной EVRA Ventil Hygiene HV 30-500-1100 без крепежа",  "HV 30-500-1100", "9926,09",  "1 шт"),
    ("Радиатор стальной EVRA Ventil Hygiene HV 30-500-1200 без крепежа",  "HV 30-500-1200", "10564,82", "1 шт"),
    ("Радиатор стальной EVRA Ventil Hygiene HV 30-500-1600 без крепежа",  "HV 30-500-1600", "13254,62", "1 шт"),
    ("Кронштейн для гигиенических радиаторов 500 мм",                     "",               "315,00",   "58 шт"),
    ("Радиатор стальной EVRA Ventil Hygiene HV 30-900-1000 без крепежа",  "HV 30-900-1000", "14329,60", "1 шт"),
    ("Радиатор стальной EVRA Ventil Hygiene HV 30-900-1200 без крепежа",  "HV 30-900-1200", "16643,91", "1 шт"),
    ("Кронштейн для гигиенических радиаторов 900 мм",                     "",               "367,50",   "4 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 21-500-1100 без крепежа",  "CV 21-500-1100", "8163,38",  "1 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 21-500-1600 без крепежа",  "CV 21-500-1600", "10779,76", "3 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 22-500-1400 без крепежа",  "CV 22-500-1400", "10859,37", "2 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 22-500-1600 без крепежа",  "CV 22-500-1600", "12180,85", "5 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 33-500-1100 без крепежа",  "CV 33-500-1100", "12540,68", "1 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 33-500-1200 без крепежа",  "CV 33-500-1200", "13396,68", "5 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 33-500-1400 без крепежа",  "CV 33-500-1400", "15086,74", "1 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 33-500-1600 без крепежа",  "CV 33-500-1600", "16961,82", "3 шт"),
    ("Кронштейн настенный для радиатора 500-21/22/33",                    "",               "212,06",   "42 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 11-500-400 без крепежа",   "CV 11-500-400",  "3584,94",  "1 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 11-500-1800 без крепежа",  "CV 11-500-1800", "8198,28",  "4 шт"),
    ("Кронштейн настенный правый для радиатора 11-го типа",               "",               "38,67",    "2 шт"),
    ("Кронштейн настенный левый для радиатора 11-го типа",                "",               "38,67",    "2 шт"),
    ("Кронштейн настенный центральный 40 мм для радиатора 11-го типа",    "",               "53,64",    "20 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 33-500-1800 без крепежа",  "CV 33-500-1800", "18681,42", "1 шт"),
    ("Кронштейн для гигиенических радиаторов 500 мм",                     "",               "315,00",   "2 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 11-500-400 без крепежа",   "CV 11-500-400",  "3584,94",  "2 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 11-500-800 без крепежа",   "CV 11-500-800",  "4850,95",  "1 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 11-500-1800 без крепежа",  "CV 11-500-1800", "8198,28",  "1 шт"),
    ("Кронштейн настенный правый для радиатора 11-го типа",               "",               "38,67",    "6 шт"),
    ("Кронштейн настенный левый для радиатора 11-го типа",                "",               "38,67",    "6 шт"),
    ("Кронштейн настенный центральный 40 мм для радиатора 11-го типа",    "",               "53,64",    "5 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 21-500-900 без крепежа",   "CV 21-500-900",  "7056,57",  "1 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 22-500-500 без крепежа",   "CV 22-500-500",  "5432,23",  "1 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 33-500-1800 без крепежа",  "CV 33-500-1800", "18681,42", "2 шт"),
    ("Кронштейн настенный для радиатора 500-21/22/33",                    "",               "212,06",   "10 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 11-400-1400 без крепежа",  "CV 11-400-1400", "5890,02",  "60 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 11-400-1600 без крепежа",  "CV 11-400-1600", "6377,03",  "135 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 11-500-500 без крепежа",   "CV 11-500-500",  "3899,52",  "1 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 11-400-1000 без крепежа",  "CV 11-400-1000", "4805,49",  "1 шт"),
    ("Кронштейн настенный правый для радиатора 11-го типа",               "",               "38,67",    "394 шт"),
    ("Кронштейн настенный левый для радиатора 11-го типа",                "",               "38,67",    "394 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 21-400-700 без крепежа",   "CV 21-400-700",  "5245,41",  "15 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 21-400-1100 без крепежа",  "CV 21-400-1100", "6964,14",  "15 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 22-400-700 без крепежа",   "CV 22-400-700",  "5739,04",  "15 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 22-400-900 без крепежа",   "CV 22-400-900",  "6673,25",  "30 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 22-400-1000 без крепежа",  "CV 22-400-1000", "7296,78",  "124 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 22-400-1100 без крепежа",  "CV 22-400-1100", "7768,39",  "240 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 22-400-1200 без крепежа",  "CV 22-400-1200", "8230,06",  "30 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 33-400-700 без крепежа",   "CV 33-400-700",  "7608,34",  "78 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 33-400-800 без крепежа",   "CV 33-400-800",  "8289,76",  "15 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 33-400-1000 без крепежа",  "CV 33-400-1000", "9832,78",  "3 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 33-600-500 без крепежа",   "CV 33-600-500",  "8049,52",  "96 шт"),
    ("Радиатор стальной EVRA Compact C 22-400-600 без крепежа",           "C 22-400-600",   "4014,79",  "1 шт"),
    ("Радиатор стальной EVRA Compact C 22-400-400 без крепежа",           "C 22-400-400",   "3069,36",  "16 шт"),
    ("Радиатор стальной EVRA Compact C 22-400-800 без крепежа",           "C 22-400-800",   "4937,72",  "16 шт"),
    ("Радиатор стальной EVRA Compact C 22-400-1000 без крепежа",          "C 22-400-1000",  "6033,98",  "32 шт"),
    ("Кронштейн настенный для радиатора 400-21/22/33",                    "",               "212,06",   "1260 шт"),
    ("Кронштейн настенный для радиатора 600-21/22/33",                    "",               "212,06",   "192 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 21-500-1600 без крепежа",  "CV 21-500-1600", "10779,76", "1 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 21-500-1800 без крепежа",  "CV 21-500-1800", "11808,79", "1 шт"),
    ("Радиатор стальной EVRA Ventil Compact CV 33-500-1400 без крепежа",  "CV 33-500-1400", "15086,74", "1 шт"),
    ("Кронштейн настенный для радиатора 500-21/22/33",                    "",               "212,06",   "7 шт"),
]

assert len(ITEMS) == 66, f"Expected 66 items, got {len(ITEMS)}"

# ── helpers ────────────────────────────────────────────────────────────────────
def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, bg="1F4E79", fg="FFFFFF", bold=True, wrap=True):
    cell.font = Font(bold=bold, color=fg, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = thin_border()

def style_data(cell, wrap=True, align="left"):
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border = thin_border()
    cell.font = Font(size=10)


# ══════════════════════════════════════════════════════════════════════════════
# Лист 1: СВОДНЫЙ (одна строка = один счёт, все позиции через ;)
# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "СВОДНЫЙ (1 строка = 1 счёт)"

# строка 1: заголовок
ws1.merge_cells("A1:I1")
ws1["A1"] = "ФАЗА 0 — Эталонные данные счетов"
ws1["A1"].font = Font(bold=True, size=13)
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 28

# строка 2: заголовки колонок
COLS_S = [
    ("A", "№",                  5),
    ("B", "Файл счёта",         45),
    ("C", "Поставщик\n(из счёта)", 22),
    ("D", "Кол-во\nпозиций",    10),
    ("E", "Наименования\n(через ;)", 60),
    ("F", "Артикулы\n(через ;)",    28),
    ("G", "Цена за ед. с НДС\n(через ;)", 28),
    ("H", "Кол-во и ед.изм.\n(через ;)", 22),
    ("I", "Родительский\nраздел", 20),
]
ws1.row_dimensions[2].height = 40
for col_letter, title, width in COLS_S:
    cell = ws1[f"{col_letter}2"]
    cell.value = title
    style_header(cell)
    ws1.column_dimensions[col_letter].width = width

# строка 3: примечание серым
ws1.merge_cells("A3:I3")
note = ws1["A3"]
note.value = (
    "⚠ Правила заполнения: "
    "разделитель = ; (точка с запятой, БЕЗ Enter внутри ячейки)  |  "
    "Артикул = код из счёта (HV 20-500-400, CV 33-500-1100 и т.д.), если нет — оставить пустым в той позиции  |  "
    "Цены — с НДС, через запятую как десятичный знак (4067,48)  |  "
    "Кол-во — число + ед.изм. через пробел (6 шт)"
)
note.font = Font(italic=True, size=9, color="444444")
note.fill = PatternFill("solid", fgColor="FFF2CC")
note.alignment = Alignment(wrap_text=True, vertical="center")
ws1.row_dimensions[3].height = 42

# строка 4: данные
names  = "; ".join(i[0] for i in ITEMS)
arts   = "; ".join(i[1] for i in ITEMS)
prices = "; ".join(i[2] for i in ITEMS)
qtys   = "; ".join(i[3] for i in ITEMS)

row4 = [
    6,
    "Оцера Радиаторы Счет на оплату № 4171 от 26.06.2025 (2).pdf",
    'ООО "ИТАС"',
    66,
    names,
    arts,
    prices,
    qtys,
    "",
]
ws1.row_dimensions[4].height = 300
for col_idx, val in enumerate(row4, start=1):
    cell = ws1.cell(row=4, column=col_idx, value=val)
    style_data(cell, wrap=True)
    if col_idx == 1:
        cell.alignment = Alignment(horizontal="center", vertical="top")
    if col_idx == 4:
        cell.alignment = Alignment(horizontal="center", vertical="top")

ws1.freeze_panes = "A4"


# ══════════════════════════════════════════════════════════════════════════════
# Лист 2: ПО ПОЗИЦИЯМ (одна строка = одна позиция, удобно для проверки)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("ПО ПОЗИЦИЯМ (для проверки)")

ws2.merge_cells("A1:H1")
ws2["A1"] = "ФАЗА 0 — Расширенный вид (одна строка = одна позиция счёта)"
ws2["A1"].font = Font(bold=True, size=13)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

COLS_P = [
    ("A", "№\nпозиции",  8),
    ("B", "Файл счёта",  45),
    ("C", "Поставщик",   22),
    ("D", "Наименование",60),
    ("E", "Артикул",     20),
    ("F", "Цена с НДС",  14),
    ("G", "Кол-во",       8),
    ("H", "Ед.изм.",      8),
]
ws2.row_dimensions[2].height = 36
for col_letter, title, width in COLS_P:
    cell = ws2[f"{col_letter}2"]
    cell.value = title
    style_header(cell, bg="375623")
    ws2.column_dimensions[col_letter].width = width

FILENAME = "Оцера Радиаторы Счет на оплату № 4171 от 26.06.2025 (2).pdf"
SUPPLIER = 'ООО "ИТАС"'

for idx, (name, art, price, qty_str) in enumerate(ITEMS, start=1):
    qty_parts = qty_str.split()
    qty_num = qty_parts[0]
    qty_unit = qty_parts[1] if len(qty_parts) > 1 else "шт"
    row_data = [idx, FILENAME, SUPPLIER, name, art, price, qty_num, qty_unit]
    r = idx + 2
    ws2.row_dimensions[r].height = 32
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws2.cell(row=r, column=col_idx, value=val)
        style_data(cell, wrap=(col_idx == 4))
        if col_idx in (1, 7):
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # zebra
        if idx % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="F5F5F5")

ws2.freeze_panes = "A3"

# ── сохранить ─────────────────────────────────────────────────────────────────
out = "scripts/phase0_sample.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Sheet 1 'СВОДНЫЙ': 1 строка = 1 счёт (формат для вашего phase0_audit.xlsx)")
print(f"Sheet 2 'ПО ПОЗИЦИЯМ': 1 строка = 1 позиция (удобно для ручной сверки)")
print(f"Позиций: {len(ITEMS)}")
