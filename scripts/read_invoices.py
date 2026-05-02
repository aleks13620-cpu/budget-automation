"""
Read all invoice files (PDF + Excel) from data/uploads/
Produce benchmark-draft.xlsx with columns A-S.
"""

import os
import re
import sys
import pdfplumber
import openpyxl
import xlrd
from openpyxl.styles import Font, PatternFill, Alignment

UPLOADS = r"c:\Users\home\vscode101\budget-automation\data\uploads"
OUTPUT  = r"c:\Users\home\vscode101\budget-automation\scripts\benchmark-draft.xlsx"
MAX_SIZE = 5 * 1024 * 1024   # 5 MB — large files are likely scanned drawings

sys.stdout.reconfigure(encoding='utf-8')


def is_copy(name):
    return "— копия" in name or "- копия" in name


def clean(s):
    return str(s).strip() if s is not None else ""


def parse_price(s):
    if s is None:
        return None
    s = (str(s)
         .replace('\xa0', '').replace(' ', '').replace(' ', '')
         .replace(',', '.').strip())
    s = re.sub(r'[^\d.]', '', s)
    try:
        v = float(s)
        return v if v > 0 else None
    except Exception:
        return None


def fmt_price(v):
    return f"{v:.2f}" if v is not None else ""


# ─── VAT detection ───────────────────────────────────────────────────────────

def detect_vat(text):
    """Return (rate_str, is_usn, is_included).
    is_included=True → unit prices in table already contain VAT.
    """
    t = text.lower()

    if re.search(r'\bусн\b|\bне облагается\b|без ндс|ндс\s*не', t):
        return "0% УСН", True, True

    # "в том числе НДС" / "в т.ч. НДС" → prices WITH VAT
    is_included = bool(re.search(
        r'в\s+том\s+числе\s+ндс|в\s*т\s*\.?\s*ч\s*\.?\s*ндс', t))
    # "стоимость/цена с учётом/учетом НДС"
    if not is_included:
        is_included = bool(re.search(
            r'(стоимость|цена)[^\n]{0,30}с\s+уч[её]том\s+ндс|с\s+уч[её]том\s+ндс', t))

    if re.search(r'20\s*%|ставка\s*20|ндс\s*20', t):
        return "20%", False, is_included
    if re.search(r'10\s*%|ставка\s*10', t):
        return "10%", False, is_included
    return "20% дефолт", False, is_included


# ─── Table parsing ───────────────────────────────────────────────────────────

HEADER_KWS = [
    'наименование', 'название', 'позиция', 'кол', 'цена',
    'сумма', 'ед', 'шт', 'артикул', 'товар',
]

# Skip rows matching these patterns at the START of the name field
SKIP_NAMES = re.compile(
    r'^(итого|всего|total|в том числе|в\s*т\s*\.?\s*ч\b|ндс\b|сумм[аы]\b|налог|'
    r'руководитель|директор|должность|бухгалтер|подпись|ответственный|'
    r'менеджер|исполнитель|гл\.?\s*бух)',
    re.IGNORECASE,
)


def find_header_row(rows):
    """Return index of header row.
    Skip rows where any single cell is longer than 80 chars (text blob, not table header).
    """
    for i, row in enumerate(rows):
        # Reject rows with a giant cell (pdfplumber sometimes folds entire page text into one cell)
        if any(len(clean(c)) > 80 for c in row if c):
            continue
        row_text = " ".join(clean(c).lower() for c in row if c)
        hits = sum(1 for kw in HEADER_KWS if kw in row_text)
        if hits >= 2:
            return i
    return None


def find_col(headers, *terms):
    """Return column index where any term is found in the header (case-insensitive).
    Columns are scanned left-to-right; first matching column wins.
    """
    for i, h in enumerate(headers):
        hl = clean(h).lower().replace('\n', ' ')
        if any(t in hl for t in terms):
            return i
    return None


def find_col_ordered(headers, *terms):
    """Like find_col but terms are tried in priority order.
    Each term is checked across ALL columns before the next term is tried.
    Use this when earlier terms should take precedence over later ones.
    """
    for term in terms:
        for i, h in enumerate(headers):
            hl = clean(h).lower().replace('\n', ' ')
            if term in hl:
                return i
    return None


def extract_items_from_table(rows, is_vat_included=False):
    """Return (items, price_col_type, comment).

    price_col_type: 'with_vat' | 'without_vat' | 'total_with_vat' | 'unclear'
    'total_with_vat' means the price cell holds a ROW TOTAL (needs /qty for unit price).
    """
    header_idx = find_header_row(rows)
    if header_idx is None:
        return [], 'unclear', "нет заголовка"

    headers = rows[header_idx]

    # Use ordered search: 'наименование' wins over 'товар' even if 'товар' appears
    # earlier in the row (e.g. "Код товара" before "Наименование товара")
    col_name = find_col_ordered(headers,
                                'наименование', 'название', 'позиция', 'описание',
                                'товар', 'услуга')
    if col_name is None:
        return [], 'unclear', "нет колонки наименование"

    col_qty  = find_col(headers,
                        'кол-во', 'кол.', 'количество', 'кол\n', 'кол во',
                        'объём', 'объем', 'кол')
    col_unit = find_col(headers,
                        'ед.', 'ед.изм', 'единица', 'ед\n', 'ед изм', 'ед.')

    # ── Price column detection (priority order) ──────────────────────────────
    # 1. Explicit unit-price-with-VAT
    col_pw = find_col(headers,
                      'цена с ндс', 'цена,\nс ндс', 'цена с\nндс',
                      'цена (с ндс)', 'цена\nс ндс', 'сумма с ндс')
    # 2. Row-TOTAL with VAT (must divide by qty)
    col_ptotal = None
    if col_pw is None:
        col_ptotal = find_col(headers, 'всего с ндс', 'итого с ндс')
    # 3. Discounted price (treat as with_vat when is_vat_included)
    col_disco = None
    if col_pw is None and col_ptotal is None:
        col_disco = find_col(headers,
                             'цена со скидкой', 'со скидкой', 'сумма со скидкой')
    # 4. Explicit unit-price-without-VAT
    col_pwo = None
    if col_pw is None and col_ptotal is None and col_disco is None:
        col_pwo = find_col(headers,
                           'цена без ндс', 'цена\nбез', 'цена, без',
                           'цена (без ндс)', 'цена\nбез ндс')
    # 5. Generic "Цена ..."
    col_pg = None
    if col_pw is None and col_ptotal is None and col_disco is None and col_pwo is None:
        col_pg = find_col(headers, 'цена\n', 'цена,', 'цена за', 'цена ')
        if col_pg is None:
            col_pg = find_col(headers, 'цена')

    # VAT sum per row (for situation 3)
    col_vsum = find_col(headers, 'ндс, руб', 'сумма ндс', 'в т.ч. ндс', 'ндс\n')

    # Assign price column and type
    if col_pw is not None:
        price_col, price_col_type = col_pw, 'with_vat'
    elif col_ptotal is not None:
        price_col, price_col_type = col_ptotal, 'total_with_vat'
    elif col_disco is not None:
        price_col = col_disco
        price_col_type = 'with_vat' if is_vat_included else 'unclear'
    elif col_pwo is not None:
        price_col, price_col_type = col_pwo, 'without_vat'
    elif col_pg is not None:
        price_col = col_pg
        price_col_type = 'with_vat' if is_vat_included else 'unclear'
    else:
        price_col, price_col_type = None, 'unclear'

    items = []
    for row in rows[header_idx + 1:]:
        if not any(clean(c) for c in row):
            continue
        name_val = clean(row[col_name]) if col_name < len(row) else ""
        if not name_val or len(name_val) < 2:
            continue
        if SKIP_NAMES.match(name_val):
            continue
        qty_raw   = clean(row[col_qty])   if col_qty   is not None and col_qty   < len(row) else ""
        unit_val  = clean(row[col_unit])  if col_unit  is not None and col_unit  < len(row) else ""
        price_raw = clean(row[price_col]) if price_col is not None and price_col < len(row) else ""
        vat_s_raw = clean(row[col_vsum])  if col_vsum  is not None and col_vsum  < len(row) else ""
        items.append({
            'name': name_val, 'qty': qty_raw, 'unit': unit_val,
            'price_raw': price_raw, 'price_col_type': price_col_type,
            'vat_sum_raw': vat_s_raw,
        })
    return items, price_col_type, None


# ─── Price calculation ───────────────────────────────────────────────────────

def calc_price_with_vat(item, vat_rate_str, is_usn, is_vat_included):
    """Return (price_str, how_label, comment)."""
    price = parse_price(item['price_raw'])
    qty   = parse_price(item['qty'])
    vat_s = parse_price(item['vat_sum_raw'])
    ptype = item['price_col_type']

    if is_usn:
        return fmt_price(price), "с НДС", ""

    if ptype == 'with_vat':
        return fmt_price(price), "с НДС", ""

    if ptype == 'total_with_vat':
        # price = row total WITH VAT, divide by qty for unit price
        if price and qty and qty > 0:
            return fmt_price(price / qty), "с НДС", ""
        return fmt_price(price), "с НДС", "нет qty для деления"

    if ptype == 'without_vat':
        if vat_s and qty and qty > 0 and price:
            return fmt_price((price * qty + vat_s) / qty), "без НДС+сумма НДС", ""
        rate = 1.10 if vat_rate_str == "10%" else 1.20
        if price:
            return fmt_price(price * rate), "без НДС+ставка", ""
        return "", "неясно", "нет цены"

    # unclear
    if price:
        rate = 1.10 if vat_rate_str == "10%" else 1.20
        comment = "НДС под вопросом" if vat_rate_str == "20% дефолт" else ""
        return fmt_price(price * rate), "без НДС+ставка", comment
    return "", "неясно", ""


def pick_three(items):
    n = len(items)
    if n == 0:  return None, None, None
    if n == 1:  return items[0], None, None
    if n == 2:  return items[0], None, items[1]
    return items[0], items[n // 2], items[-1]


def item_fields(item, vat_rate, is_usn, is_vat_included):
    if item is None:
        return "", "", "", ""
    price_str, _, _ = calc_price_with_vat(item, vat_rate, is_usn, is_vat_included)
    return item['name'], price_str, item['qty'], item['unit']


# ─── File processors ─────────────────────────────────────────────────────────

def process_pdf(path):
    try:
        with pdfplumber.open(path) as pdf:
            text_parts, all_tables = [], []
            for pg in pdf.pages:
                t = pg.extract_text()
                if t:
                    text_parts.append(t)
                for tbl in pg.extract_tables():
                    if tbl:
                        all_tables.append(tbl)
            text = "\n".join(text_parts)
    except Exception as e:
        return {'status': 'нечитаемый', 'count': '', 'items': [],
                'vat_rate': '20% дефолт', 'is_usn': False, 'is_vat_included': False,
                'price_how': 'неясно', 'comment': f"ошибка: {e}"}

    if not text or len(text.strip()) < 10:
        return {'status': 'нечитаемый', 'count': 0, 'items': [],
                'vat_rate': '20% дефолт', 'is_usn': False, 'is_vat_included': False,
                'price_how': 'неясно', 'comment': "нет текста (скан?)"}

    vat_rate, is_usn, is_vat_included = detect_vat(text)
    all_items, comments = [], []
    table_count = 0

    for tbl in all_tables:
        items, ptype, issue = extract_items_from_table(tbl, is_vat_included)
        if issue and issue != "нет заголовка":
            comments.append(issue)
        if items:
            all_items.extend(items)
            table_count += 1

    if table_count > 1:
        comments.append(f"несколько таблиц ({table_count})")
    if not all_tables and text:
        comments.append("текст без таблиц")

    price_how = "неясно"
    if all_items:
        _, price_how, extra = calc_price_with_vat(
            all_items[0], vat_rate, is_usn, is_vat_included)
        if extra and extra not in comments:
            comments.append(extra)

    status = 'читаемый' if all_items else ('под вопросом' if text else 'нечитаемый')
    return {
        'status': status, 'count': len(all_items), 'items': all_items,
        'vat_rate': vat_rate, 'is_usn': is_usn, 'is_vat_included': is_vat_included,
        'price_how': price_how,
        'comment': "; ".join(dict.fromkeys(c for c in comments if c)),
    }


def process_excel(path):
    ext = os.path.splitext(path)[1].lower()
    sheets, err = [], None

    if ext == '.xlsx':
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            for ws in wb.worksheets:
                rows = [
                    [str(c) if c is not None else "" for c in row]
                    for row in ws.iter_rows(values_only=True)
                ]
                sheets.append((ws.title, rows))
        except Exception as e:
            err = str(e)
    else:
        try:
            wb = xlrd.open_workbook(path)
            for sh in wb.sheets():
                rows = []
                for r in range(sh.nrows):
                    row = []
                    for c in range(sh.ncols):
                        cell = sh.cell(r, c)
                        if cell.ctype == 2:    # XL_CELL_NUMBER
                            v = cell.value
                            row.append(str(int(v)) if v == int(v) else str(v))
                        elif cell.ctype == 3:  # XL_CELL_DATE
                            row.append(str(cell.value))
                        else:
                            row.append(str(cell.value).strip())
                    rows.append(row)
                sheets.append((sh.name, rows))
        except Exception as e:
            err = str(e)

    if err or not sheets:
        return {'status': 'нечитаемый', 'count': '', 'items': [],
                'vat_rate': '20% дефолт', 'is_usn': False, 'is_vat_included': False,
                'price_how': 'неясно', 'comment': err or 'пустой файл'}

    all_text = " ".join(
        " ".join(c for row in rows for c in row)
        for _, rows in sheets
    )
    vat_rate, is_usn, is_vat_included = detect_vat(all_text)
    all_items, comments = [], []

    if len(sheets) > 1:
        comments.append(f"{len(sheets)} листов")

    for sheet_name, rows in sheets:
        items, ptype, issue = extract_items_from_table(rows, is_vat_included)
        if issue and issue not in ('нет заголовка', 'нет колонки наименование'):
            comments.append(f"{sheet_name}: {issue}")
        if items:
            all_items.extend(items)

    price_how = "неясно"
    if all_items:
        _, price_how, extra = calc_price_with_vat(
            all_items[0], vat_rate, is_usn, is_vat_included)
        if extra and extra not in comments:
            comments.append(extra)

    return {
        'status': 'читаемый' if all_items else 'под вопросом',
        'count': len(all_items), 'items': all_items,
        'vat_rate': vat_rate, 'is_usn': is_usn, 'is_vat_included': is_vat_included,
        'price_how': price_how,
        'comment': "; ".join(dict.fromkeys(c for c in comments if c)),
    }


# ─── Build Excel output ──────────────────────────────────────────────────────

HEADERS = [
    'A: Файл', 'B: Статус', 'C: Кол-во строк товаров',
    'D: Наим. (начало)', 'E: Цена с НДС (начало)', 'F: Кол-во (начало)', 'G: Ед.изм (начало)',
    'H: Наим. (середина)', 'I: Цена с НДС (середина)', 'J: Кол-во (середина)', 'K: Ед.изм (середина)',
    'L: Наим. (конец)', 'M: Цена с НДС (конец)', 'N: Кол-во (конец)', 'O: Ед.изм (конец)',
    'P: Ставка НДС', 'Q: Как указана цена', 'R: Комментарий', 'S: Проверка',
]


def make_row(fname, data):
    items = data['items']
    vat   = data['vat_rate']
    usn   = data['is_usn']
    incl  = data['is_vat_included']
    first, mid, last = pick_three(items)
    d, e, f, g = item_fields(first, vat, usn, incl)
    h, i, j, k = item_fields(mid,   vat, usn, incl)
    l, m, n, o = item_fields(last,  vat, usn, incl)
    return [
        fname, data['status'], data['count'],
        d, e, f, g, h, i, j, k, l, m, n, o,
        vat, data['price_how'], data.get('comment', ''), '',
    ]


def main():
    all_files = sorted(os.listdir(UPLOADS))
    unique = [
        f for f in all_files
        if not is_copy(f)
        and os.path.splitext(f)[1].lower() in ('.pdf', '.xlsx', '.xls')
    ]
    print(f"Files to process: {len(unique)}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Эталон"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    for fname in unique:
        fpath = os.path.join(UPLOADS, fname)
        size  = os.path.getsize(fpath)
        ext   = os.path.splitext(fname)[1].lower()
        print(f"  {fname!r}  {size // 1024} KB", flush=True)

        if size > MAX_SIZE:
            data = {
                'status': 'нечитаемый', 'count': '', 'items': [],
                'vat_rate': '20% дефолт', 'is_usn': False, 'is_vat_included': False,
                'price_how': 'неясно',
                'comment': f'{size // 1024 // 1024} МБ — вероятно скан/чертёж',
            }
        else:
            try:
                data = process_pdf(fpath) if ext == '.pdf' else process_excel(fpath)
            except Exception as ex:
                data = {
                    'status': 'нечитаемый', 'count': '', 'items': [],
                    'vat_rate': '20% дефолт', 'is_usn': False, 'is_vat_included': False,
                    'price_how': 'неясно', 'comment': str(ex),
                }

        row = make_row(fname, data)
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        print(f"    status={data['status']!r}  count={data['count']}"
              f"  vat={data['vat_rate']!r}  incl={data['is_vat_included']}"
              f"  how={data['price_how']!r}  comment={data.get('comment', '')!r}")

    widths = [42, 12, 10, 48, 16, 10, 10, 48, 16, 10, 10, 48, 16, 10, 10, 15, 22, 45, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(OUTPUT)
    print(f"\nDone → {OUTPUT}")


if __name__ == '__main__':
    main()
