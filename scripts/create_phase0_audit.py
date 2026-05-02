"""
Генерирует эталонный Excel-лист фазы 0 для Ивана и Сергея.
Каждый счёт — одна строка; столбцы = данные для ручной фиксации.
"""

import os
from pathlib import Path
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

UPLOADS_DIR = Path(r"C:\Users\home\vscode101\budget-automation\data\uploads")
OUTPUT_FILE = Path(r"C:\Users\home\vscode101\budget-automation\scripts\phase0_audit.xlsx")

# --- collect files, drop copies -----------------------------------------
SKIP_MARKERS = ("— копия", "- копия", "—копия", "-копия")

def is_copy(name: str) -> bool:
    stem = Path(name).stem
    return any(m in stem for m in SKIP_MARKERS)

all_files = sorted(UPLOADS_DIR.iterdir(), key=lambda p: p.name.lower())
SKIP_FILES = {".gitkeep", "Thumbs.db", "desktop.ini"}
invoices = [
    f for f in all_files
    if f.is_file()
    and not is_copy(f.name)
    and f.name not in SKIP_FILES
    and not f.name.startswith(".")
]

# --- colour palette ---------------------------------------------------------
CLR_HEADER_BG  = "1F3864"   # тёмно-синий
CLR_HEADER_FG  = "FFFFFF"
CLR_SECTION_BG = "D6E4F0"   # светло-голубой для групп проверок
CLR_ODD_ROW    = "EAF2FB"
CLR_EVEN_ROW   = "FFFFFF"
CLR_BORDER     = "B0C4DE"

thin = Side(style="thin", color=CLR_BORDER)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_style(ws, row, col, value, bg=CLR_HEADER_BG, fg=CLR_HEADER_FG,
              size=10, bold=True, wrap=True, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
    cell.border = border
    return cell

def data_style(ws, row, col, value="", bg=CLR_EVEN_ROW, align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(size=9, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=True)
    cell.border = border
    return cell

# --- build workbook ---------------------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Фаза 0 — Эталон"

# ── ROW 1 : title ───────────────────────────────────────────────────────────
ws.merge_cells("A1:K1")
title = ws["A1"]
title.value = "ФАЗА 0 — Эталонные данные счетов (заполняется вручную)"
title.font = Font(bold=True, size=13, color=CLR_HEADER_FG, name="Calibri")
title.fill = PatternFill("solid", fgColor=CLR_HEADER_BG)
title.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# ── ROW 2 : group labels ─────────────────────────────────────────────────────
GROUP_BG = "2E75B6"
ws.merge_cells("A2:C2");  hdr_style(ws, 2, 1, "ИДЕНТИФИКАЦИЯ",       bg=GROUP_BG, size=9)
ws.merge_cells("D2:D2");  hdr_style(ws, 2, 4, "ОШИБКА №2",           bg="C00000", size=9)
ws.merge_cells("E2:F2");  hdr_style(ws, 2, 5, "ОШИБКА №3",           bg="C00000", size=9)
ws.merge_cells("G2:H2");  hdr_style(ws, 2, 7, "ОШИБКА №1",           bg="C00000", size=9)
ws.merge_cells("I2:J2");  hdr_style(ws, 2, 9, "БАЗОВАЯ ТОЧНОСТЬ",    bg=GROUP_BG, size=9)
ws.merge_cells("K2:K2");  hdr_style(ws, 2, 11,"ГРУППИРОВКА",         bg=GROUP_BG, size=9)
ws.row_dimensions[2].height = 18

# ── ROW 3 : column headers ───────────────────────────────────────────────────
COLS = [
    # (ширина, заголовок)
    (5,  "№"),
    (38, "Файл счёта"),
    (18, "Поставщик\n(из счёта)"),
    (10, "Кол-во\nпозиций\n(эталон)"),   # error #2
    (30, "Наименования позиций\n(все, через ;\nесли искажений нет — «ОК»)"),  # error #3
    (14, "Статус\nнаименований\n(ОК / Ошибка)"),
    (14, "Цена за ед.\nс НДС\n(все через ;)"),  # error #1
    (14, "Статус цен\n(ОК / Обнуление)"),
    (20, "Кол-во и\nед.изм.\n(все через ;)"),
    (14, "Статус кол-ва\n(ОК / Ошибка)"),
    (22, "Родительский раздел\n(если есть)"),
]
for ci, (width, label) in enumerate(COLS, start=1):
    hdr_style(ws, 3, ci, label, bg="2E75B6", size=9)
    ws.column_dimensions[get_column_letter(ci)].width = width
ws.row_dimensions[3].height = 48

# ── DATA ROWS ───────────────────────────────────────────────────────────────
for ri, fpath in enumerate(invoices, start=1):
    row = ri + 3
    bg = CLR_ODD_ROW if ri % 2 == 1 else CLR_EVEN_ROW

    ext = fpath.suffix.upper().lstrip(".")
    data_style(ws, row, 1, ri,         bg=bg, align="center")
    data_style(ws, row, 2, fpath.name, bg=bg)
    for ci in range(3, 12):
        data_style(ws, row, ci, "", bg=bg)

    ws.row_dimensions[row].height = 32

# ── freeze panes, auto-filter ───────────────────────────────────────────────
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:K{3 + len(invoices)}"

# ── instruction sheet ───────────────────────────────────────────────────────
wi = wb.create_sheet("Инструкция")
instructions = [
    ("ИНСТРУКЦИЯ ДЛЯ ПРОВЕРЯЮЩИХ (Иван / Сергей)", True, 13),
    ("", False, 10),
    ("Цель:", True, 11),
    ("Заполнить лист «Фаза 0 — Эталон» вручную по оригинальным счетам.", False, 10),
    ("Эти данные станут эталоном для автоматической проверки парсера.", False, 10),
    ("", False, 10),
    ("Столбцы:", True, 11),
    ("№ — порядковый номер", False, 10),
    ("Файл счёта — имя файла в папке uploads", False, 10),
    ("Поставщик — название поставщика из счёта", False, 10),
    ("Кол-во позиций (эталон) — сколько строк-товаров реально в счёте (число)", False, 10),
    ("Наименования позиций — все наименования товаров/работ через точку с запятой ;", False, 10),
    ("  Если наименования парсер передаёт без искажений — пишите «ОК»", False, 10),
    ("Статус наименований — ОК / Ошибка", False, 10),
    ("Цена за ед. с НДС — все цены через ; (порядок совпадает с наименованиями)", False, 10),
    ("Статус цен — ОК / Обнуление (если хоть одна цена стала 0)", False, 10),
    ("Кол-во и ед.изм. — все значения через ; (например: 5 шт; 2 м²; 10 компл)", False, 10),
    ("Статус кол-ва — ОК / Ошибка", False, 10),
    ("Родительский раздел — заголовок раздела/группы если счёт структурирован по разделам", False, 10),
    ("", False, 10),
    ("Приоритет проверки (ошибки, которые уже были выявлены):", True, 11),
    ("  Ошибка №1 — Обнуление цен (столбцы G–H)", False, 10),
    ("  Ошибка №2 — Потеря строк: кол-во позиций (столбец D)", False, 10),
    ("  Ошибка №3 — Искажение наименований (столбцы E–F)", False, 10),
]

for row_i, (text, bold, size) in enumerate(instructions, start=1):
    cell = wi.cell(row=row_i, column=1, value=text)
    cell.font = Font(bold=bold, size=size, name="Calibri")
    cell.alignment = Alignment(wrap_text=True)

wi.column_dimensions["A"].width = 80

# ── save ─────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"OK: {OUTPUT_FILE}")
print(f"  invoices (no copies): {len(invoices)}")
for f in invoices:
    print(f"    {f.name}")
