"""Debug: print raw table headers for each PDF/Excel file."""
import os, sys
import pdfplumber
import openpyxl
import xlrd

UPLOADS = r'c:\Users\home\vscode101\budget-automation\data\uploads'

def is_copy(n): return '— копия' in n or '- копия' in n

files = sorted(f for f in os.listdir(UPLOADS)
               if not is_copy(f) and os.path.splitext(f)[1].lower() in ('.pdf','.xlsx','.xls'))

sys.stdout.reconfigure(encoding='utf-8')

for fname in files:
    fpath = os.path.join(UPLOADS, fname)
    size = os.path.getsize(fpath)
    ext = os.path.splitext(fname)[1].lower()
    print(f"\n=== {fname} ({size//1024} KB) ===")

    if size > 5*1024*1024:
        print("  [skipped - too large]")
        continue

    if ext == '.pdf':
        try:
            with pdfplumber.open(fpath) as pdf:
                tbl_count = 0
                for pi, page in enumerate(pdf.pages):
                    for tbl in page.extract_tables():
                        if not tbl: continue
                        tbl_count += 1
                        if tbl_count <= 3:
                            print(f"  Table {tbl_count} (page {pi+1}), rows={len(tbl)}")
                            for ri, row in enumerate(tbl[:6]):
                                print(f"    row{ri}: {[str(c)[:40] if c else None for c in row]}")
                if tbl_count == 0:
                    print("  [no tables]")
        except Exception as e:
            print(f"  ERROR: {e}")

    elif ext == '.xlsx':
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
            for ws in wb.worksheets:
                print(f"  Sheet: {ws.title!r}")
                rows = list(ws.iter_rows(values_only=True, max_row=20))
                for ri, row in enumerate(rows):
                    if any(v is not None for v in row):
                        print(f"    row{ri}: {[str(v)[:30] if v is not None else None for v in row[:12]]}")
        except Exception as e:
            print(f"  ERROR: {e}")

    elif ext == '.xls':
        try:
            wb = xlrd.open_workbook(fpath)
            for sh in wb.sheets():
                print(f"  Sheet: {sh.name!r}")
                for ri in range(min(20, sh.nrows)):
                    row = [str(sh.cell(ri, c).value)[:30] for c in range(sh.ncols)]
                    if any(v.strip() for v in row):
                        print(f"    row{ri}: {row[:12]}")
        except Exception as e:
            print(f"  ERROR: {e}")
